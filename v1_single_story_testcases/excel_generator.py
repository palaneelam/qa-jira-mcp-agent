import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FOLDER = Path("output")


def ensure_output_folder():
    OUTPUT_FOLDER.mkdir(exist_ok=True)


def format_excel(file_path):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path)

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="D9EAF7",
                end_color="D9EAF7",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(max_length + 5, 70)

    workbook.save(file_path)


def normalize_test_cases(test_cases):
    if test_cases is None:
        return []

    if isinstance(test_cases, list):
        return test_cases

    if isinstance(test_cases, dict):
        if "test_cases" in test_cases:
            return test_cases["test_cases"]

        if "data" in test_cases:
            return test_cases["data"]

        if "error" in test_cases:
            print("\nAI returned error:")
            print(test_cases)
            return []

        return [test_cases]

    return [{
        "Raw_Response": str(test_cases)
    }]


def export_test_cases(test_cases, issue_key):
    ensure_output_folder()

    normalized_data = normalize_test_cases(test_cases)

    print("\nNormalized Test Cases:")
    print(normalized_data)

    if not normalized_data:
        print("\nNo test case data available. Excel will not be generated.")
        return

    file_path = OUTPUT_FOLDER / f"{issue_key}_TestCases.xlsx"

    df = pd.DataFrame(normalized_data)

    print("\nDataFrame Preview:")
    print(df.head())

    df.to_excel(
        file_path,
        sheet_name="TestCases",
        index=False
    )

    format_excel(file_path)

    print(f"\nExcel created: {file_path}")